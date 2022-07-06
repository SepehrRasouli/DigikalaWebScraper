# Python scraper to scrape any given subject from www.digikala.com
# and write them Into excel
from bs4 import BeautifulSoup
from bs4.element import SoupStrainer
import requests
from openpyxl import load_workbook
import re
import os
import datetime

now = datetime.datetime.now()
time_sheet = ""
meta = ""

DIGIKALA_FILTERS = [
    "&only_plus=1&",
    "&only_fresh=1&",
    "&has_ship_by_seller=1&",
    "&has_jet_delivery=1&",
    "&has_selling_stock=1&",
    "&has_ready_to_shipment=1&",
    "&seller_condition[0]=digikala&",
    "&seller_condition[1]=official&",
    "&seller_condition[2]=trusted&",
    "&seller_condition[3]=roosta&",
    "&sortby=7&",
    "&sortby=22&",
    "&sortby=4&",
    "&sortby=1&",
    "&sortby=20&",
    "&sortby=21&",
    "&sortby=25&",
]

date_to_write = []


def cleanizer(filename):
    """Cleans up extracted data."""
    print("Cleaning up more...")
    global date_to_write
    date_to_write = []
    if filename == "extracted_prices.txt":
        # To handle تومان better
        with open(filename, "r", encoding="utf-8") as f:
            context = f.readlines()
            context = [x.strip() for x in context]
        empty_file(filename)
        for element in context:
            if element == "تومان" or element == "":
                continue
            if element == "None":
                date_to_write.append("ناموجود")
                continue

            else:
                date_to_write.append(element + " تومان ")
        os.remove(filename)
        del context

    elif filename == "extracted_discount_values.txt":
        with open(filename, "r", encoding="utf-8") as f:
            context = f.readlines()
            context = [x.strip() for x in context]
        empty_file(filename)
        for element in context:
            if element:
                date_to_write.append(element)

            else:
                continue
        os.remove(filename)
        del context
    elif filename == "extracted_stars.txt":
        with open(filename, "r", encoding="utf-8") as f:
            context = f.readlines()
            context = [x.strip() for x in context]
        empty_file(filename)
        with open(filename, "a", encoding="utf-8") as f:
            for element in context:
                if element.startswith("(") or element == "":
                    continue
                else:
                    date_to_write.append(element + "\n")

        del context
        os.remove(filename)

    else:
        with open(filename, "r", encoding="utf-8") as f:
            context = f.readlines()
            context = [x.strip() for x in context]
        empty_file(filename)
        for element in context:
            if element == "فروش ویژه" or element == "Ad" or element == "":
                continue
            else:
                date_to_write.append(element)
        os.remove(filename)
        del context


def writer(excel_file):
    global meta, date_to_write, time_sheet
    """Writes the data to an Excel file."""
    # checkfilevalidity
    print("Writing data...")
    if os.path.isfile(excel_file):
        try:
            wb = load_workbook(filename=excel_file)
            if time_sheet not in wb.sheetnames:
                wb.create_sheet(time_sheet)
            else:
                pass
            ws = wb[time_sheet]
            colnum = 0
            if meta == "discount_values":
                colnum = 1
            elif meta == "stars":
                colnum = 2
            elif meta == "names":
                colnum = 3
            else:
                colnum = 4
            i = 1
            for element in date_to_write:
                ws.cell(row=i, column=colnum, value=element)
                i += 1
            wb.save(filename=excel_file)
            wb.close()
            del date_to_write

        except PermissionError:
            print("Please close the Excel file and try again.")
            del date_to_write
            main()
    else:
        print("The Excel file doesn't exist!")
        del date_to_write
        main()


def empty_file(filename):
    """Removes all context of a file."""
    with open(filename, "w") as temp:
        temp.write(" ")


def extract_data(subject, pages_limit, selected_filters):
    """Gets data from digikala search."""

    empty_file("pages_result.txt")
    for num in range(1, pages_limit):
        try:
            with open("pages_result.txt", "a", encoding="utf-8") as f:
                meta = "&sortby=22" if num > 1 else ""
                r = requests.get(
                    f"https://www.digikala.com/search/?{''.join(selected_filters)}q={subject}&pageno={num}{meta}"
                )
                print(
                    f"https://www.digikala.com/search/?{''.join(selected_filters)}q={subject}&pageno={num}{meta}"
                )
                f.write(str(r.text) + "\n")
        except ConnectionError:
            print("Got An Connection Error. Please Check your internet Connection")


def extractor(filename):
    """Extracts the wanted context."""
    print("Extracting data...")
    global meta
    if "products_discounts_result.txt" == filename:
        meta = "discount_values"
    elif "products_stars_result.txt" == filename:
        meta = "stars"
    elif "products_names_result.txt" == filename:
        meta = "names"
    else:
        meta = "prices"
    extracted_file = f"extracted_{meta}.txt"

    with open(extracted_file, "a", encoding="utf-8") as result:
        with open(filename, "r", encoding="utf-8") as mainFile:
            lines = mainFile.readlines()
            regexed = re.sub(r"<[^>]*>", "\n", "".join(lines))
            for char in regexed:
                if char in [",", "]", "[", "[,"]:
                    continue
                else:
                    result.write(char)

    print("Cleaning up...")
    os.remove(filename)
    cleanizer(extracted_file)


def products_discount_values_scraper():
    print("Extracting discount values...")
    with open("pages_result.txt", "r", encoding="utf-8") as f:
        with open("products_discounts_result.txt", "a+", encoding="utf-8") as result:
            only_discount_box = SoupStrainer(
                "div", {"class": "c-product-box__row c-product-box__row--price"}
            )
            context = f.readlines()
            soup = BeautifulSoup(
                "".join(context), "html.parser", parse_only=only_discount_box
            )
            containers = soup.find_all(
                "div", class_="c-product-box__row c-product-box__row--price"
            )
            for container in containers:
                if "c-price__discount-oval" in str(container):
                    result.write(
                        str(container.find("div", class_="c-price__discount-oval"))
                        + "\n"
                    )
                    continue
                else:
                    result.write("%۰" + "\n")
                    continue

            del context, soup, containers

    extractor("products_discounts_result.txt")


def products_stars_scraper():
    print("Extracting Stars...")
    with open("pages_result.txt", "r", encoding="utf-8") as f:
        with open("products_stars_result.txt", "a+", encoding="utf-8") as result:
            context = f.readlines()
            engagement = SoupStrainer("div", {"class": "c-product-box__content"})
            soup = BeautifulSoup("".join(context), "html.parser", parse_only=engagement)
            containers = soup.find_all("div", class_="c-product-box__content")
            for container in containers:
                if "c-product-box__engagement-rating" in str(container):
                    result.write(
                        str(
                            container.find(
                                "div", class_="c-product-box__engagement-rating"
                            )
                        )
                        + "\n"
                    )
                    continue
                else:
                    result.write("۰.۰" + "\n")
                    continue

            del context, soup, containers

    extractor("products_stars_result.txt")


def products_names_scraper():
    print("Extracting product names...")
    with open("pages_result.txt", "r", encoding="utf-8") as f:
        with open("products_names_result.txt", "a+", encoding="utf-8") as result:
            context = f.readlines()
            productNames = SoupStrainer("a", {"class": "js-product-url"})
            soup = BeautifulSoup("".join(context), "html.parser", parse_only=productNames)
            productName = soup.find_all("a", class_="js-product-url")
            if productName and productName != " ":
                result.write(str(productName))

            else:
                result.write("Not Found.")
            del context, soup, productName

    extractor("products_names_result.txt")


def products_prices_scraper():
    print("Extracting prices...")
    with open("pages_result.txt", "r", encoding="utf-8") as f:
        with open("products_prices_result.txt", "a+", encoding="utf-8") as result:
            context = f.readlines()
            Prices = SoupStrainer(
                "div", {"class": "c-product-box__row c-product-box__row--price"}
            )
            soup = BeautifulSoup("".join(context), "html.parser", parse_only=Prices)
            containers = soup.find_all(
                "div", class_="c-product-box__row c-product-box__row--price"
            )
            for container in containers:
                if container.find(
                    "div",
                    class_="c-price__value c-price__value--plp js-plp-product-card-price",
                ):
                    result.write(
                        str(
                            container.find(
                                "div",
                                class_="c-price__value c-price__value--plp js-plp-product-card-price",
                            ).find("div", class_="c-price__value-wrapper")
                        )
                        + "\n"
                    )
                else:
                    result.write(
                        str(container.find("div", class_="c-price__value-wrapper"))
                        + "\n"
                    )
            del context, soup, containers

    extractor("products_prices_result.txt")


def check(user_input):
    checklist = []
    alphabet = "abcdefghijklmnopqrstuvwxyz,ABCDEFGHIJKLMNOPQRSTUVWXYZ!@#$%^&*()/\«»<>_+"
    for char in alphabet:
        if char in user_input:
            checklist.append(False)
        checklist.append(True)

    if all(checklist):
        return True


def main():
    global excel_file, now, time_sheet
    use_current_data = 0
    # Greet to user and take scraping options
    print("Digikala web scraper v2 (optimized)")
    print("Hello, user! I can scrape Digikala and write data to excel.")
    print("*" * 10)
    selected_options = []
    selected_filters = []
    subject = input("Subject > ")
    # TODO: Make sure that `pages_limit`` is an integer

    pages_limit = input("How many pages? > ")

    if not check(pages_limit):
        print("Invalid number!")
        main()

    if not pages_limit:
        print("Invalid number!")
        main()
    try:
        pages_limit = int(pages_limit) + 1  # Because of range function\
    except ValueError as error:
        print(f"Got an {error}. Did you give too many numbers?")
        main()

    excel_file = input("Excel file > ")

    if excel_file == " " or "":
        main()
    if excel_file.endswith(".xlsx"):
        pass
    else:
        excel_file += ".xlsx"

    print("*" * 10)
    print(
        "Scarping options : *Select One Or More Options and seperate them with spaces* "
    )
    print(
        """
    1- Products names 2- Products prices
    3- Products discount values 4- Products stars"""
    )
    print("*" * 10)

    user_input = input("> ")

    result = check(user_input)

    # Add selected scarping options to list

    if result and user_input != "":
        for number in set(user_input.split()):
            if int(number) > 4:
                print("Invalid option!")
                main()

            else:
                if int(number) == 1:
                    selected_options.append("products_names_scraper()")

                if int(number) == 2:
                    selected_options.append("products_prices_scraper()")

                if int(number) == 3:
                    selected_options.append("products_discount_values_scraper()")

                if int(number) == 4:
                    selected_options.append("products_stars_scraper()")

    else:
        print("Alphabet character or ',' detected!")
        main()

    if os.path.isfile("pages_result.txt"):
        print("Use current Digikala web data? [Y/n] ")
        yon = input("> ")
        if yon.lower().startswith("y"):
            use_current_data = 1

        elif yon.lower().startswith("n"):
            use_current_data = 0

        else:
            print("Wrong choice!")
            main()
    else:
        use_current_data = 0
        pass

    print("Digikala filters: ")
    print("*" * 10)
    print("Select one or more options and seperate them with spaces: ")
    print(
        """
    ** Choosing too much options or bad options might affect Digikala searchs.**
    1- Only DigiPlus                            10- Seller (indigenous seller)
    2- Only Supermarkets                        11- Bestselling
    3- Ship by seller                           12- Most relevant
    4- Fast delivery                            13- Most visited
    5- Only avalaibles                          14- Newests
    6- Only avalaibles in DigiKala's store      15- Cheapest
    7- Seller (DigiKala)                        16- Most expensives
    8- Seller (officials)                       17- Fastest post
    9- Seller (trusted)                         18- None
    """
    )
    user_input = input("> ")
    result = check(user_input)

    if result == True:
        print("Extracting data from Digikala...")
        user_input = set(user_input.split())
        for i in user_input:
            i = int(i)
            if i == 18:
                break

        else:
            for selected_option in user_input:
                if int(selected_option) < 18:
                    selected_option = int(selected_option) - 1
                    selected_filters.append(
                        DIGIKALA_FILTERS[selected_option]
                    )  # Add Elements By Index
                    continue
                else:
                    continue
    else:
        print("Invalid choice!")
        main()

    if use_current_data == 0:
        extract_data(subject, pages_limit, selected_filters)
    else:
        pass

    time_sheet = (
        f"{now.year} {now.day} {now.month} {now.hour} {now.minute} {now.second}"
    )
    for option in selected_options:
        exec(option)
        writer(excel_file)

    print("All done!\n\n")
    print("*" * 10)
    main()


main()
